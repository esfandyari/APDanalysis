import sys
import openpyxl as oxl
import numpy as np
import matplotlib.pyplot as plt
from scipy import signal
import peakutils
from scipy import interpolate

########################### peak filtering #########################
def peak_filter(peak_h,peak_id,minimum):
    p_id_new = []
    for i in range(len(peak_id)):
        if peak_h[i] > minimum:
            p_id_new = np.append(p_id_new,int(peak_id[i]))
            p_id_new = p_id_new.astype(int)
    return p_id_new

########################## peak modifier ###########################

def peak_modifier(p_id,data,ran):
    peaks_ids_new = np.copy(p_id)
    for i in range(len(p_id)):
        max_p = data[1][p_id[i]]
        if p_id[i] > ran+1 and p_id[i] < len(data[1][:])-ran-1:
            for j in range(ran):
                if data[1][p_id[i]+j] > max_p :
                    peaks_ids_new[i] = p_id[i]+j
                    max_p = data[1][p_id[i]+j]
            for j in range(ran):
                if data[1][p_id[i]-j] > max_p:
                    peaks_ids_new[i] = p_id[i]-j
                    max_p = data[1][p_id[i]-j]
        if p_id[i] < ran:
            for j in range(ran):
                if data[1][p_id[i]+j] > max_p :
                    peaks_ids_new[i] = p_id[i]+j
                    max_p = data[1][p_id[i]+j]
            for j in range(p_id[i]):
                if data[1][p_id[i]-j] > max_p:
                    peaks_ids_new[i] = p_id[i]-j
                    max_p = data[1][p_id[i]-j]
        if p_id[i] > len(data[1][:])-ran-1:
            for j in range(len(data[1][:])-p_id[i]-1):
                if data[1][p_id[i]+j] > max_p :
                    peaks_ids_new[i] = p_id[i]+j
                    max_p = data[1][p_id[i]+j]
            for j in range(ran):                
                if data[1][p_id[i]-j] > max_p:
                    peaks_ids_new[i] = p_id[i]-j
                    max_p = data[1][p_id[i]-j]
    return peaks_ids_new

#####################################righ and left###########################


def peak_r_l(data,baseline,p_h,peak_id,delta,tol,data_f,max_it):
    
    p_x_rl = np.zeros([2,len(p_h)])
    
    for i in range(len(p_h)):
        base_ex         = interpolate.interp1d(data[0][:],baseline+p_h[i])
        p_x_rl[0][i]    = intersect(data[0][peak_id[i]],base_ex,data_f,delta,tol,max_it,"R")
        p_x_rl[1][i]    = intersect(data[0][peak_id[i]],base_ex,data_f,delta,tol,max_it,"L")
        
    return p_x_rl;


# finds the intersection between "func1" and "func2" from the right side.
    # "init"    = initial point
    # "delta"   = the initial steps
    # "tol"     = error tolerance
    # "max_it"  = maximum iterations
    # "direction" = determine the direction of root finding, either Right ("R") or Left ("L")

def intersect(init,baseline,data,delta,tol,max_it,direction):
    point_x     = init
    error       = abs(data(point_x) - baseline(point_x))
    point_x_temp   = point_x
    counter     = 0
    #print("the peak is equal to %f and the baseline is equal to %f" % (data(init),baseline(init)))
    while error > tol and counter <= max_it:
       # print("The error is now %f and delta is %f" %( error, delta))
        counter += 1
        if direction == "R":
            point_x_temp += delta
        else:
            point_x_temp -= delta
            
        if data(point_x_temp) - baseline(point_x_temp) >= 0:
           # print("Did not hit! The baseline is %f and the data is %f " %(baseline(point_x_temp) , data(point_x_temp)))
            point_x = point_x_temp
            error = abs(data(point_x) - baseline(point_x))
            
        else:
            #print("This is a hit! The baseline is %f and the data is %f!" %(baseline(point_x_temp) , data(point_x_temp)))        
            delta = delta/2.0
            
    return point_x


##########################import data from excel sheet#########################

def importer(excel_name, number,t_label,x_label):
    filename    = excel_name +str(number) # the file name
    data        = oxl.load_workbook(filename,read_only=True)   #loading the file read-only
    data_excel  = data['Sheet1']                        #enter into the desired workshee
    t_label = str(t_label)
    x_label = str(x_label)
    c_count = data_excel.max_column     #number of column
    r_count = data_excel .max_row       #number of row
    data = np.zeros((c_count,r_count))  #row data
    for i in range(1,c_count+1):
        for j in range(1,r_count+1):
            if i == 1:
                c_label = 'A'
            if i == 2:
                c_label = 'B'
            element_label = c_label+ str(j)
            data[i-1][j-1] = data_excel[element_label].value
    return data

#############################PEAK HEIGHT############################
def peak_height(peaks_id,base_f,data):
    p_h = np.zeros(len(peaks_id))
    for i in range(len(peaks_id)):
        p_h[i] = data[1][peaks_id[i]] - base_f(data[0][peaks_id[i]])
    return p_h

#############################Data Filter############################
def data_subtract_baseline(data,baseline):
    data_new = np.copy(data)
    for i in range(len(data[0][:])):
         data_new[1][i] = data[1][i] - baseline(data[0][i])
    return data_new
############################### APD 90 ##############################

def peak_h_90(p_h):
    p_h_90 = np.copy(p_h)
    for i in range(len(p_h)):
        p_h_90[i] = p_h[i] * 0.1
    return p_h_90

############################### APD 50 ##############################

def peak_h_50(p_h):
    p_h_50 = np.copy(p_h)
    for i in range(len(p_h)):
        p_h_50[i] = p_h[i] * 0.5
    return p_h_50

############################### APD 70 ##############################

def peak_h_70(p_h):
    p_h_70 = np.copy(p_h)
    for i in range(len(p_h)):
        p_h_70[i] = p_h[i] * 0.3
    return p_h_70

############################### APD 20 ##############################

def peak_h_20(p_h):
    p_h_20 = np.copy(p_h)
    for i in range(len(p_h)):
        p_h_20[i] = p_h[i] * 0.8
    return p_h_20

##################################RR calc ##############################

def rr_calc(data,peak_id):
    RR = np.zeros([len(peak_id) -1])
    for i in range(len(RR)):
        RR[i] = data[0][[peak_id[i+1]]] - data[0][[peak_id[i]]]
    return RR

################################build baseline################################

def baseline_builder(var,data):
    baseline = np.zeros(len(data[1]))
    for i in range(len(data[1])):
        baseline[i] = var
    return baseline

##############################peak-split-point##################################

def peak_split_point(data,peak_id,RR):
	RR_half 	= RR/2.0
	peak_s_p 	= np.zeros(len(RR))
	for i in range(len(RR)):
		peak_s_p[i] = data[0][peak_id[i]]+RR_half[i]
	return peak_s_p
##############################peak data generator############################

def peak_data_gen(peak_s_p,data,data_f,time_lapse):
    part = len(peak_s_p)+1
    t = time_lapse
    peak_data = np.zeros([part,len(data[1])])
    peak_s_p_n = np.insert(peak_s_p,0,data[0][0])
    peak_s_p_n = np.append(peak_s_p_n, data[0][-1])
    section =np.zeros(part)
    for i in range(part):
        section[i]=int((peak_s_p_n[i+1]-peak_s_p_n[i])/ t)
    section = section.astype(int)
    for j in range(part):
        for i in range(section[j]):
            peak_data[j][i] = data_f(peak_s_p_n[j]+(t * i) )
    return peak_data

#################################data average################################

def data_averager(peak_data,peak_location):
    section = len(peak_data)
    roll_loc = np.max(peak_location)
    for i in range(len(peak_location)):
        peak_data[i] = np.roll(peak_data[i],roll_loc-peak_location[i])
    data_average = np.zeros(len(peak_data[0]))
    for j in range(section):
        for i in range(len(peak_data[0])):
            data_average[i] += (peak_data[j][i]/section)
    return data_average

####################################peak locator#########################################

def peak_locator(peak_data):
    location = np.zeros(len(peak_data),)
    for i in range(len(peak_data)):
        peaks = signal.find_peaks_cwt(peak_data[i], np.arange(5,10))
        maximum = peaks[0]
        for j in range(len(peaks)):
            if peak_data[i][peaks[j]]> peak_data[i][maximum]:
                maximum = peaks[j]
        location[i] = maximum
    return location


######################################plot##################################################

def plotting(data_x, data_y):
	plt.plot(data_x,data_y)
	plt.show()
	
	
